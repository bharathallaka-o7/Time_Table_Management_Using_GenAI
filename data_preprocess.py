import os
import pandas as pd
import openpyxl
from openpyxl.utils.cell import range_boundaries
import numpy as np

def preprocess_excel(file_path, output_path=None):
    """
    Comprehensive Excel sheet preprocessing function with improved merged cell handling
    
    Parameters:
    -----------
    file_path : str
        Path to the input Excel file
    output_path : str, optional
        Path to save the processed Excel file. 
        If None, generates a path in the same directory
    
    Returns:
    --------
    pd.DataFrame: Processed dataframe
    """
    # If no output path is provided, generate one
    if output_path is None:
        directory, filename = os.path.split(file_path)
        filename_base, ext = os.path.splitext(filename)
        output_path = os.path.join(directory, f"{filename_base}{ext}")
    
    # Load workbook and sheet
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    
    # Step 1: Identify and report merged cells
    merged_cell_ranges = list(sheet.merged_cells.ranges)
    print(f"Merged Cell Ranges Found: {len(merged_cell_ranges)}")
    
    # Step 2: Demerge and fill cells - Improved method
    for merge_range in merged_cell_ranges:
        # Extract boundaries of the merged cell range
        min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))
        
        # Get the value from the top-left cell in the merged range
        # Use data_only=True when loading workbook to get actual cell values
        value = sheet.cell(row=min_row, column=min_col).value
        
        # Unmerge the cells first
        sheet.unmerge_cells(str(merge_range))
        
        # Fill all cells in the merged range with this value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.value = value
    
    # Temporary save to process with pandas
    temp_output_path = os.path.join(os.path.dirname(file_path), 'temp_demerged.xlsx')
    workbook.save(temp_output_path)
    
    # Step 3: Convert to pandas DataFrame for advanced processing
    df = pd.read_excel(temp_output_path)
    
    # Remove temporary file
    os.remove(temp_output_path)
    
    # Remove completely empty rows and columns
    df.dropna(how='all', axis=0, inplace=True)
    df.dropna(how='all', axis=1, inplace=True)
    
    # Reset index after dropping rows
    df.reset_index(drop=True, inplace=True)
    
    # Step 4: Clean and standardize column names
    df.columns = [str(col).strip().lower().replace(' ', '_') for col in df.columns]
    
    # Step 5: Handle potential data type inconsistencies
    # Example: Convert numeric columns to appropriate types
    for col in df.columns:
        try:
            df[col] = pd.to_numeric(df[col], errors='ignore')
        except:
            pass
    
    # Step 6: Remove any remaining completely empty rows
    df.dropna(how='all', inplace=True)
    
    # Step 7: Save processed data
    # Save as Excel
    df.to_excel(output_path, index=False)
   
    
    print(f"Processed file saved: {output_path}")
   
    return df

# Example usage
if __name__ == "__main__":
    input_file = 'Modified Data\ECE\ECE_Timetable.xlsx'
    processed_df = preprocess_excel(input_file)
    
    # Optional: Display first few rows and basic info
    print("\nProcessed DataFrame Info:")
    print(processed_df.info())
    print("\nFirst few rows:")
    print(processed_df.head())